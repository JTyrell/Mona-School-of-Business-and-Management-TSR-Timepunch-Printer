import openpyxl
from datetime import date, datetime
import calendar

def process_autofill(
    file_path: str,
    output_path: str,
    year: int,
    month: int,
    sheets_to_fill: list,
    standard_schedule: dict,
    mid_month_schedule: dict
):
    """
    standard_schedule: {"days": [0, 1, 2], "start": "09:00 AM", "end": "05:00 PM", "description": "Helpdesk"}
    mid_month_schedule: {"enabled": bool, "start_date": "YYYY-MM-DD", "days": [...], "start": "...", "end": "...", "description": "Helpdesk"}
    """
    
    wb = openpyxl.load_workbook(file_path)
    
    # Generate dates for the month
    num_days = calendar.monthrange(year, month)[1]
    
    mid_month_date = None
    if mid_month_schedule and mid_month_schedule.get("enabled"):
        try:
            mid_month_date = datetime.strptime(mid_month_schedule["start_date"], "%Y-%m-%d").date()
        except ValueError:
            pass

    for sheet_name in sheets_to_fill:
        # User might type "Sheet1, Sheet2", ensure no trailing spaces
        sheet_name = sheet_name.strip()
        if sheet_name not in wb.sheetnames:
            continue
            
        sheet = wb[sheet_name]
        
        # Find header row and column indices
        header_row_idx = None
        col_map = {}
        
        # Scan first 20 rows for headers
        for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            if row is None:
                continue
            
            row_str = [str(cell).upper().strip() if cell else "" for cell in row]
            
            if any("DATE" in c for c in row_str) and any("START" in c for c in row_str):
                header_row_idx = r_idx
                for c_idx, cell_val in enumerate(row_str, start=1):
                    if "DATE" in cell_val:
                        col_map["date"] = c_idx
                    elif "DESCRIPTION" in cell_val:
                        col_map["description"] = c_idx
                    elif "START" in cell_val:
                        col_map["start"] = c_idx
                    elif "END" in cell_val:
                        col_map["end"] = c_idx
                break
                
        if not header_row_idx or "date" not in col_map:
            # Could not find header in this sheet, skip
            continue
            
        # Start filling from the row below the header
        current_row = header_row_idx + 1
        
        for day in range(1, num_days + 1):
            current_date = date(year, month, day)
            weekday = current_date.weekday() # 0 = Monday, 6 = Sunday
            
            # Determine which schedule applies
            active_schedule = standard_schedule
            if mid_month_date and current_date >= mid_month_date:
                active_schedule = mid_month_schedule
                
            active_day_conf = active_schedule.get(str(weekday))
            if active_day_conf and active_day_conf.get("enabled"):
                # We need to fill this row
                # Format date properly
                # Write proper datetime to retain Excel formatting
                date_obj = datetime.combine(current_date, datetime.min.time())
                desc_str = active_schedule.get("description", "Helpdesk")
                start_str = active_day_conf.get("start", "")
                end_str = active_day_conf.get("end", "")
                
                # Write to cells (Intentionally omitting HOURS and LINE TOTAL to preserve formulas)
                for col_name, val in [("date", date_obj), ("description", desc_str), ("start", start_str), ("end", end_str)]:
                    if col_name in col_map:
                        try:
                            sheet.cell(row=current_row, column=col_map[col_name]).value = val
                        except AttributeError:
                            pass
                
                current_row += 1

        # Clear any leftover rows from a previous month's template
        # Check the next 31 rows to ensure no ghost data remains
        for r in range(current_row, current_row + 32):
            if "date" in col_map:
                try:
                    date_cell = sheet.cell(row=r, column=col_map["date"])
                    date_val = date_cell.value
                    
                    if date_val is None or str(date_val).strip() == "":
                        # Reached the end of the populated rows
                        break
                        
                    date_str = str(date_val).strip().lower()
                    if "total" in date_str or "signature" in date_str or "summary" in date_str:
                        # Reached the totals/signature section
                        break

                    # It's a leftover data row, clear it
                    for col_name in ["date", "description", "start", "end"]:
                        if col_name in col_map:
                            try:
                                sheet.cell(row=r, column=col_map[col_name]).value = None
                            except AttributeError:
                                pass
                except AttributeError:
                    pass

    wb.save(output_path)
    return True
